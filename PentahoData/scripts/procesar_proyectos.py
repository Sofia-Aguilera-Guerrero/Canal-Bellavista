import sys
sys.stdout.reconfigure(encoding='utf-8')
# procesar_proyectos.py
# -*- coding: utf-8 -*-

"""
Script unificado para:
  1) Asegurar librerías (pandas, openpyxl).
  2) Insertar columna ESTATUS después de 'Monto UF' (y rellenar 'Construido' en filas 5..29) EN MEMORIA.
  3) Detectar y consolidar todas las tablas válidas de la Hoja4 en una gran tabla.
  4) Guardar un ÚNICO archivo final en C:\PentahoData\output\.

Rutas esperadas (creadas previamente):
  C:\PentahoData\input\Proyectos-LAPTOP-Q10B3SCV.xlsx
  C:\PentahoData\output\
  C:\PentahoData\scripts\procesar_proyectos.py  (este archivo)
"""

import os, sys, subprocess, io, re
from datetime import datetime

# ========= PARÁMETROS DE RUTAS =========
INPUT_XLSX = r"C:\PentahoData\input\Proyectos-LAPTOP-Q1OB3SCV.xlsx"
OUTPUT_DIR = r"C:\PentahoData\output"
SHEET_NAME = "Hoja4"

# ========= PARÁMETROS DE LA 1ª TABLA =========
# (fila 4 son encabezados; filas 5..29 son los datos de esa primera tabla)
ROW_HEADERS_1B = 4   # fila de encabezados en Excel (1-based)
FILL_FROM_1B   = 5   # primera fila con datos de la 1ª tabla (1-based)
FILL_TO_1B     = 29  # última fila con datos de la 1ª tabla (1-based)

# ========= BÚSQUEDA DE TABLAS =========
VENTANA            = 15  # filas hacia abajo para buscar el próximo encabezado
LOOKAHEAD_HEADER   = 2   # filas extra para completar el header a 6 columnas

# ========= 0) ASEGURAR LIBRERÍAS =========
def ensure_packages():
    need = []
    try:
        import pandas  # noqa
    except Exception:
        need.append("pandas")
    try:
        import openpyxl  # noqa
    except Exception:
        need.append("openpyxl")
    if need:
        print(f"Instalando paquetes: {', '.join(need)} ...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", *need])
        print("Paquetes instalados.\n")

ensure_packages()

import pandas as pd
from openpyxl import load_workbook


# ========= UTILIDADES =========
def norm(s: str) -> str:
    """Normaliza texto para comparar encabezados."""
    s = "" if s is None else str(s).strip().lower()
    s = (s.replace("á", "a").replace("é", "e").replace("í", "i")
           .replace("ó", "o").replace("ú", "u").replace("ñ", "n"))
    s = re.sub(r"[\.]", "", s)
    s = re.sub(r"\s+", " ", s)
    return s

VARIANTS = {
    "concurso":        {"concurso"},
    "proyecto":        {"proyecto"},
    "construido_por":  {"construido por", "construidor por", "empresa", "contruido por"},
    "costo_total_uf":  {"costo total (uf)", "costo total uf", "costo total", "monto uf"},
    "lugar_ref":       {"lugar ref", "lugar referencia"},
    "estatus":         {"estatus"},
}

ORDERED_COLS = ["CONCURSO", "PROYECTO", "CONSTRUIDO POR",
                "COSTO TOTAL (UF)", "LUGAR REF.", "ESTATUS"]


def map_row(serie: pd.Series):
    """Devuelve dict lógico→índice de columna si encuentra títulos."""
    rn = [norm(x) for x in serie.tolist()]
    m = {}
    for j, val in enumerate(rn):
        for k, vs in VARIANTS.items():
            if val in vs and k not in m:
                m[k] = j
    return m

def is_min_header(m: dict) -> bool:
    """Encabezado mínimo requerido para considerar bloque válido."""
    return all(k in m for k in ("concurso", "proyecto", "construido_por"))

def is_full_header(m: dict) -> bool:
    """Encabezado completo (las 6 columnas)."""
    return all(k in m for k in ("concurso","proyecto","construido_por","costo_total_uf","lugar_ref","estatus"))

def s(x) -> str:
    return "" if x is None or (isinstance(x, float) and pd.isna(x)) else str(x).strip()


# ========= 1) PASO 2 EN MEMORIA: insertar ESTATUS y rellenar filas 5..29 =========
def paso2_insertar_estatus_en_memoria(path_xlsx: str, sheet: str,
                                      row_headers_1b: int, r_from_1b: int, r_to_1b: int) -> io.BytesIO:
    if not os.path.isfile(path_xlsx):
        raise FileNotFoundError(f"No existe el archivo de entrada: {path_xlsx}")

    wb = load_workbook(path_xlsx)
    if sheet not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet}' en el libro.")

    ws = wb[sheet]

    # Encontrar columna "Monto UF" en la fila de encabezados
    col_ref = None
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=row_headers_1b, column=c).value
        if isinstance(val, str) and val.strip().lower() == "monto uf":
            col_ref = c
            break
    if col_ref is None:
        raise ValueError("No encontré la columna 'Monto UF' en la fila de encabezados.")

    # Insertar nueva columna a la derecha con cabecera "ESTATUS"
    col_estatus = col_ref + 1
    ws.insert_cols(col_estatus)
    ws.cell(row=row_headers_1b, column=col_estatus, value="ESTATUS")

    # Rellenar "Construido" solo en la primera tabla (filas 5..29)
    for r in range(r_from_1b, r_to_1b + 1):
        ws.cell(row=r, column=col_estatus, value="Construido")

    # Guardar libro modificado a un buffer en memoria (sin archivo intermedio)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ========= 2) PASO 3: consolidar tablas detectando encabezados =========
def complete_header_map(raw_df: pd.DataFrame, row_idx: int, lookahead_rows: int = 2):
    """Desde row_idx (encabezado mínimo) mira algunas filas extra para completar a 6 columnas."""
    n_rows = raw_df.shape[0]
    combined = {}
    header_row = None
    end = min(row_idx + 1 + lookahead_rows, n_rows)
    for r in range(row_idx, end):
        m = map_row(raw_df.iloc[r, :])
        if r == row_idx and not is_min_header(m):
            return None, None
        if r == row_idx:
            header_row = r
        for k, j in m.items():
            combined.setdefault(k, j)
        if is_full_header(combined):
            return header_row, combined
    return None, None

def find_header_in_window(raw_df: pd.DataFrame, start_row: int,
                          window: int = VENTANA, lookahead_rows: int = LOOKAHEAD_HEADER):
    """Busca un encabezado completo (con lookahead) entre start_row y start_row+window."""
    n_rows = raw_df.shape[0]
    end_row = min(start_row + window, n_rows)
    for r in range(start_row, end_row):
        m_base = map_row(raw_df.iloc[r, :])
        if is_min_header(m_base):
            h_row, full_map = complete_header_map(raw_df, r, lookahead_rows)
            if full_map is not None:
                return h_row, full_map
    return None, None

def paso3_consolidar(desde_buffer: io.BytesIO, sheet: str) -> pd.DataFrame:
    raw = pd.read_excel(desde_buffer, sheet_name=sheet, header=None, dtype=str)
    n_rows = raw.shape[0]
    registros = []
    pos = 0
    bloque = 0

    while True:
        h_row, h_map = find_header_in_window(raw, pos, VENTANA, LOOKAHEAD_HEADER)
        if h_map is None:
            print("No se encontraron más encabezados.")
            break

        bloque += 1
        print(f"🔎 Encabezado #{bloque} detectado en fila {h_row+1} (Excel 1-based).")

        # Leer filas de datos desde la siguiente
        r = h_row + 1
        filas_tomadas = 0
        while r < n_rows:
            fila = raw.iloc[r, :]

            # ¿Nuevo encabezado mínimo? → terminar bloque y continuar desde allí
            if is_min_header(map_row(fila)):
                pos = r
                break

            # ¿CONCURSO vacío? → fin de bloque (no agregamos esta fila)
            j_con = h_map.get("concurso", None)
            v_con = s(fila.iloc[j_con]) if j_con is not None and j_con < len(fila) else ""
            if v_con == "":
                pos = r + 1
                print(f" CONCURSO vacío en fila {r+1}. Fin del bloque #{bloque}.")
                break

            def get_val(k):
                j = h_map.get(k, None)
                return s(fila.iloc[j]) if j is not None and j < len(fila) else ""

            registros.append({
                "CONCURSO":         get_val("concurso"),
                "PROYECTO":         get_val("proyecto"),
                "CONSTRUIDO POR":   get_val("construido_por"),
                "COSTO TOTAL (UF)": get_val("costo_total_uf"),
                "LUGAR REF.":       get_val("lugar_ref"),
                "ESTATUS":          get_val("estatus"),
            })
            filas_tomadas += 1
            r += 1

        if filas_tomadas:
            print(f"Bloque #{bloque}: {filas_tomadas} filas agregadas.")

        # Protección anti-loop
        if pos <= h_row:
            pos = h_row + 1

        if pos >= n_rows:
            break

    gran = pd.DataFrame(registros, columns=ORDERED_COLS)
    # Limpieza mínima
    gran.replace({"nan": None, "None": None, "": None}, inplace=True)
    return gran


# ========= MAIN =========
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Paso 2 en memoria (insertar ESTATUS y rellenar filas 5..29)
    buf = paso2_insertar_estatus_en_memoria(
        INPUT_XLSX, SHEET_NAME, ROW_HEADERS_1B, FILL_FROM_1B, FILL_TO_1B
    )
    print("ESTATUS insertado (después de 'Monto UF') y 'Construido' rellenado 5..29 (en memoria).")

    # Paso 3: consolidar
    gran = paso3_consolidar(buf, SHEET_NAME)

    # Vista previa antes de guardar
    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", None)
    print("\nVista previa de las primeras 50 filas:\n")
    print(gran.head(50))

    # Guardar archivo fijo (sin fecha, siempre se sobrescribe)
    out_fixed = os.path.join(OUTPUT_DIR, "gran_tabla_proyectos.xlsx")

    # Si ya existe, eliminarlo para evitar errores si estaba abierto
    try:
        os.remove(out_fixed)
    except Exception:
        pass

    gran.to_excel(out_fixed, index=False)

    print(f"\n[OK] Filas totales (sin encabezado): {len(gran)}")
    print(f"[OK] Archivo final guardado (sobrescrito): {out_fixed}")


if __name__ == "__main__":
    main()


